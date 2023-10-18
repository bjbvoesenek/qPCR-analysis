#%%
# -*- coding: utf-8 -*-
"""
Script for automating the analysis of qPCR data
"""

__version_info__ = ('2','2','1')
__version__ = '.'.join(__version_info__)

#%% Check user input
import sys
# input to script from web page: python qpcr_analysis.py --input Analysis.xlsx --genes GAPDH Bactin --controls Control1 Control2 FLB240

import argparse
parser = argparse.ArgumentParser(description='Automated analysis of qPCR data exported from LinRegPCR')
parser.add_argument(
    '--version', action='version',
    version='%(prog)s {version}'.format(version=__version__))
parser.add_argument(
    '--input', action='store', nargs=1, required=True, metavar='<input file>',
    help='an .xlsx saved from LinRegPCR')
parser.add_argument(
    '--genes', action='store', nargs='+', metavar='<gene>',
    help='a list of housekeeping genes, each separated by a space')
parser.add_argument(
    '--controls', action='store', nargs='+', metavar='<control>',
    help='a list of control cell lines, each separated by a space')

args = vars(parser.parse_args())
extract_data = False # Should we just extract the genes and cell lines?

if args['input'] == None:
    print('Error: No input file given. Provide the Excel file to be analyzed using --input.\n')
    sys.exit(1)
else:
    if args['genes'] == None and args['controls'] == None:
        extract_data = True
        input_file = args['input'][0]
    elif args['genes'] == None or args['controls'] == None:
        print('Error: Wrong number of arguments. Provide the Excel file to be analysed only, or provide the Excel file to be analyzed together with housekeeping genes (--genes) and control lines (--controls).\n')
        sys.exit(2)
    else:
        input_file = args['input'][0]
        housekeeping_genes = args['genes']
        control = args['controls']

import os
# Check if the script has access to write files to the current directory
if os.access(os.getcwd(), os.W_OK) == False:
    sys.exit('Error: No rights to write files to the current directory.\n')

# Check if analysis is already performed in the current directory
if os.path.isdir("Data") or os.path.isdir("Figures"):
    sys.exit('Error: An analysis has already been performed in this folder. Please go to a different folder and run the script there.\n')


#%% Determine whether experiment is done in LinRegPCR or BioRad

import pandas as pd
# Check if analysis excel file exist
if not input_file.lower().endswith(('.xlsx')):
    print('Error: The input file is not an .xlsx file. Check if you provided the right file.\n')
    sys.exit(3)

# Check if the right sheet exist
from openpyxl import load_workbook
user_wb = load_workbook(input_file, read_only = True)

if 'Cq' in user_wb.sheetnames:
    data = pd.read_excel(input_file, sheet_name='Cq')
    if pd.Series(['Fluor', 'Target', 'Content']).isin(data.columns).all(): # Second check, should return True if BioRad was used
        qPCR_system = 'BioRad'
        data = data.reset_index(drop=True)
    else:
        sys.exit('Error: Not sure whether this is a BioRad or LinRegPCR output file. Check the names of the sheets and columns and try again.\nSee manual for more information.\n')
elif 'Data' in user_wb.sheetnames:
    data = pd.read_excel(input_file, sheet_name='Data')
    if pd.Series(['input.txt', 'Text']).isin(data.columns).all(): # Second check, should return True if LinRegPCR was used
        qPCR_system = 'LinRegPCR'
    else:
        sys.exit('Error: Not sure whether this is a BioRad or LinRegPCR output file. Check the names of the sheets and columns and try again.\nSee manual for more information.\n')
else:
    print('Error: Naming of sheets not recognized. Make sure the sheets in your excel workbook are named correctly.\nSee the manual for more information.\n')
    sys.exit(4)

#%% For the BioRad system, remove wells without sample

if qPCR_system == 'BioRad':
    # Find rows in data table containing the information of empty wells
    empty_wells = []
    for i in range(0,data.shape[0]):
        if str(data.loc[i,'Sample']) == 'nan':
           empty_wells.append(data.loc[i,'Well'])

    # Delete these rows from the data
    for i in range(0,data.shape[0]):
        if data.loc[i, 'Well'] in empty_wells:
            data = data.drop([i])
    data = data.reset_index(drop=True)

    # Delete these wells from melting curves sheet (if melting curves are provided)
    if 'Melting curves' in user_wb.sheetnames:
        df_melting = pd.read_excel(input_file, sheet_name='Melting curves')
        df_melting = df_melting.drop(columns=[col for col in df_melting if col in empty_wells])


#%% Find unique primers and cell lines

import matplotlib.pyplot as plt
import math
from natsort import natsorted, index_natsorted, order_by_index

if qPCR_system == 'LinRegPCR':
    if 'input.txt' in data:
        del data['input.txt']

    sample_names = data['Text']

elif qPCR_system == 'BioRad':
    for i in range(0,data.shape[0]):
        temp_sample_name = [data.loc[i, 'Sample'], data.loc[i,'Target']]
        data.loc[i,'Sample names'] = '_'.join(str(e) for e in temp_sample_name)

    sample_names = data['Sample names']

index = index_natsorted(sample_names)
sample_names_df = sample_names.to_frame()

# Find unique primers and samples
primer_names = sample_names_df.copy()
primer_names = primer_names.iloc[:,0]
cell_lines = sample_names_df.copy()
cell_lines = cell_lines.iloc[:,0]

for n in range(0,len(primer_names)):
    primer_names[n] = primer_names[n].partition('_')[2]
    cell_lines[n] = cell_lines[n].partition('_')[0]

unique_primers = primer_names.unique()
unique_cell_lines = cell_lines.unique()

# Convert ndarray to list
unique_primers = unique_primers.tolist()
unique_cell_lines = unique_cell_lines.tolist()

# Remove MQ from control line options
unique_cell_lines_MQ_removed = unique_cell_lines.copy()

water_controls = ['h2o', 'H2O', 'mq', 'MQ']
for item in water_controls:
    if item in unique_cell_lines_MQ_removed: unique_cell_lines_MQ_removed.remove(item)

#%% Store names of cell lines and primers in .txt files if user only provides Excel sheet

if extract_data:
    # Check if analysis is already performed in the current directory
    if os.path.isdir("Input"):
        sys.exit('Error: An analysis has already been performed in this folder. Please go to a different folder and run the script there.\n')

    # Create directory so store information about the user input
    os.mkdir('Input')

    # Save names of used cell lines in .txt file
    sorted_unique_cell_lines_MQ_removed = natsorted(unique_cell_lines_MQ_removed)
    with open("Input/Cell_lines.txt", "w") as txt_file:
        for line in sorted_unique_cell_lines_MQ_removed:
            txt_file.write(line + "\n")

    # Save names of used primers in .txt file
    sorted_unique_primers = natsorted(unique_primers)
    with open("Input/Genes.txt", "w") as txt_file:
        for line in sorted_unique_primers:
            txt_file.write(line + "\n")

    print(
        "Successfully stored all cell lines and genes from the input to text files.\n" +
        "Please select your housekeeping genes and control cell lines and use --genes and --controls to pass them to this program.\n" +
        "E.g.,\n" +
        "python3 " + __file__[len(os.getcwd())+1:] + " --input " + args['input'][0] + " --genes " + sorted_unique_primers[0] + " --controls " + sorted_unique_cell_lines_MQ_removed[0] + "\n"
    )
    # Exit script. User now has to select housekeeping genes and control cell lines in web-interface
    sys.exit(0)



# Check if the given values actually match.
for gene in housekeeping_genes:
    if gene.upper() not in map(str.upper, unique_primers):
        print("Error: The provided housekeeping gene '" + gene + "' could not be found in your data. Check if it's named correctly.\n")
        sys.exit(5)

for cell_line in control:
    if cell_line.upper() not in map(str.upper, unique_cell_lines):
        print("Error: The provided control '" + cell_line + "' could not be found in your data. Check if it's named correctly.\n")
        sys.exit(6)

#%% Plot qPCR amplification plots

# Create directories to store plots and data
os.mkdir('Figures')
os.mkdir('Data')

x = range(0, data.shape[1] - 1)

size_subplots = data.shape[0]
size_subplots = math.sqrt(size_subplots)
size_subplots = math.ceil(size_subplots)

nr_primersets = len(unique_primers)
nr_samples = len(unique_cell_lines)
nr_replicates = len(sample_names) / nr_samples / nr_primersets
nr_replicates = int(nr_replicates)

rows = nr_samples
columns = nr_primersets * nr_replicates

plt.figure(figsize = (5*columns,4 * rows))
plt.subplots_adjust(hspace = 0.5, wspace = 0.3)
#plt.suptitle('Melting curves', fontsize = 100)

subplot_number = 1
name_counter = 0

max_yvalue = data.max(numeric_only=True)
max_yvalue = max_yvalue.max() * 1.05

# Only plot qPCR plost if using LinRegPCR system (BioRad does not provide this data)
if qPCR_system == 'LinRegPCR':
    df_for_plotting = data.copy()
    if 'Text' in df_for_plotting: del df_for_plotting['Text']

    for i in range(len(index)):
        ax = plt.subplot(rows, columns, subplot_number)
        subplot_number += 1
        y = df_for_plotting.loc[index[i]]
        plt.plot(x,y)
        plt.ylim(0,max_yvalue)
        ax.set_title(sample_names[index[i]], fontsize = 15)
        name_counter += 1

    plt.savefig('Figures/Amplification_plots_sorted.pdf', bbox_inches='tight')

#%% Plot melting curves

if 'Melting curves' in user_wb.sheetnames:
    if qPCR_system == 'LinRegPCR': # In case of BioRad, this data is already loaded
        df_melting = pd.read_excel(input_file, sheet_name='Melting curves')

        for col in df_melting.columns:
            if col.startswith('Unnamed'):
                del df_melting[col]

        x_melting = df_melting[df_melting.columns[0]]

        for col in df_melting.columns:
            if col.startswith('X'):
                del df_melting[col]
    elif qPCR_system == 'BioRad':
        x_melting = df_melting['Temperature']
        del df_melting['Temperature']

    plt.figure(figsize = (5*columns, 4*rows))
    plt.subplots_adjust(hspace = 0.5, wspace = 0.3)
    #plt.suptitle('Melting curves', fontsize = 100)

    subplot_number = 1
    name_counter = 0
    index_number = 0

    y_max = max(df_melting.max()) * 1.1

    for i in range(len(index)):
        ax = plt.subplot(rows, columns, subplot_number)
        subplot_number += 1
        y = df_melting[df_melting.columns[index[i]]]
        plt.plot(x_melting,y)
        plt.ylim(-0.5,y_max)
        ax.set_title(sample_names[index[i]], fontsize = 15)
        name_counter += 1
        index_number += 1

    plt.savefig('Figures/Melting_curves_sorted.pdf', bbox_inches='tight')
else:
    print('Sheet [Melting curves] not found in your input file. Continuing to the next step...')

#%% calculate average Ct per condition

if qPCR_system == 'LinRegPCR':
    if 'Data_compact' not in user_wb.sheetnames:
        print('Error: Sheet [Data_compact] not found in your input file. Make sure the sheets in your excel workbook are named correctly.\nSee the manual for more information.\n')
        sys.exit(7)

    df_compact = pd.read_excel(input_file, sheet_name='Data_compact')

    # Depending on LinRegPCR version, Cq values are in different columns with different names
    # Find column containing Cq values
    df_compact_columns = df_compact.columns.tolist()
    for name in df_compact_columns:
        if name.startswith('Chemistry'):
            Cq_column_name = name
    Cq_index = df_compact_columns.index(Cq_column_name)

    df_Ct = df_compact.iloc[3:3+len(sample_names), Cq_index]
    df_Ct = df_Ct.to_frame()
    df_Ct = df_Ct.reset_index()
    df_Ct['Sample'] = sample_names
    del df_Ct['index']
    df_Ct = df_Ct.rename(columns={Cq_column_name : 'Ct value', 'Sample' : 'Sample'})
    df_Ct['Ct value'] = df_Ct['Ct value'].astype(float)

elif qPCR_system == 'BioRad':
    df_Ct = data[['Cq', 'Sample names']].copy()
    df_Ct = df_Ct.rename(columns={'Cq' : 'Ct value', 'Sample names' : 'Sample'})

# Remove Mq/H2O samples, as those are empty
for i in range(0,df_Ct.shape[0]):
    if df_Ct.loc[i,'Sample'].partition('_')[0] in water_controls:
        df_Ct = df_Ct.drop([i])
df_Ct = df_Ct.reset_index(drop=True)
unique_cell_lines = unique_cell_lines_MQ_removed.copy()
nr_samples = len(unique_cell_lines)

# Make df with replicates on one row
replicates_sorted = pd.DataFrame(index=range(nr_samples * nr_primersets),columns=range(nr_replicates + 1))
row_counter = 0

# Sort data, replicates in the same row
for i in range(0, len(unique_cell_lines)):
    for j in range(0, len(unique_primers)):
        replicate_name = [str(unique_cell_lines[i]), str(unique_primers[j])]
        replicate_name = '_'.join(replicate_name)
        replicates_sorted.iloc[row_counter, 0] = replicate_name
        replicate_df_temp = df_Ct[df_Ct['Sample'] == replicate_name]
        replicates_sorted.iloc[row_counter, 1:(2+nr_replicates-1)] = replicate_df_temp['Ct value'].tolist()
        row_counter += 1

## Remove outliers (SD > 0.2)!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!?!?!?!!?!?!?!?!??!??!!?!?!?!?!?!?!?!?!?!??!!?!?!!!
# outlier_threshold = 1 # times SD

# df_outliers_removed = replicates_sorted.copy()

# for i in range(0,nr_samples * nr_primersets):
#     # SD per row
#     temp_sd = df_outliers_removed.iloc[i,1:1+nr_replicates].std()
#     temp_mean = df_outliers_removed.iloc[i,1:1+nr_replicates].mean()
#     temp_max = temp_mean + temp_sd * outlier_threshold
#     temp_min = temp_mean - temp_sd * outlier_threshold
#     # Check every replicate for outliers
#     for j in range(1,1+nr_replicates):
#         if df_outliers_removed.iloc[i,j] > temp_max or df_outliers_removed.iloc[i,j] < temp_min:
#             df_outliers_removed.iloc[i,j] = None


#%% Plot Bargraph

# Determine size subplots
size_subplots = len(unique_primers)
size_subplots = math.sqrt(size_subplots)
size_subplots = math.ceil(size_subplots)

subplot_number = 1

color_list = [''] * nr_samples
avg_Ct_df = pd.DataFrame()

# Initialize figure (increase figsize if plots don't fit)
plt.figure(figsize = (5*size_subplots,5*size_subplots))
plt.subplots_adjust(hspace = 0.5, wspace = 0.3)

# Loop over unique primers and plot Ct values as bar graph
for i in range(0,len(unique_primers)):
    ax = plt.subplot(size_subplots, size_subplots, subplot_number)
    subplot_number += 1

    temp_df = replicates_sorted[replicates_sorted[0].str.partition('_')[2] == unique_primers[i]]
    temp_values = temp_df.iloc[:,1:1+nr_replicates].mean(axis=1).tolist()
    temp_samples = temp_df.iloc[:,0].tolist()
    for j in range(0,len(temp_samples)):
        temp_samples[j] = temp_samples[j].replace('_' + unique_primers[i], "")

    # Store Ct values per primer in dataframe (later save to excel)
    avg_Ct_df[unique_primers[i]] = temp_values

    for k in range(0,len(temp_samples)):
        if temp_samples[k] in control:      # Control lines will have gray bars
            color_list[k] = 'gray'
        else:                               # Investigated lines will have blue bars
            color_list[k] = 'royalblue'

    ax.bar(temp_samples, temp_values, color=color_list, alpha=0.7)
    ax.tick_params(axis='x', labelrotation=90)
    ax.set_ylim(0,40)
    ax.set_title(unique_primers[i])

    for i in range(len(unique_cell_lines)):
       ax.scatter([i] * nr_replicates, temp_df.iloc[i,1:1+nr_replicates].values.tolist(), marker='o', c='k', s=5)


plt.savefig('Figures/Average_Ct_bargraph.pdf', bbox_inches='tight')

avg_Ct_df.index = temp_samples
avg_Ct_df.to_excel("Data/Average_Ct_values.xlsx")

#%% Calculate relative expression, not normalized (2^-(dCt)))

from statistics import mean

# Initialize dataframe and check if housekeeping genes are in the dataframe
dCt_df = pd.DataFrame()

# Store column names. For analysis, it does not matter whether upper or lower case housekeeping genes is used.
original_col_names = avg_Ct_df.columns.copy()
analysis_col_names = avg_Ct_df.columns.str.upper()
unique_primers_analysis = unique_primers.copy()
unique_primers_analysis = [x.upper() for x in unique_primers_analysis]

avg_Ct_df.columns = analysis_col_names

# Subtract housekeeping gene Ct from primer Ct in every row/sample (delta Ct)
housekeeping_genes_upper = [s.upper() for s in housekeeping_genes]

for index, row in avg_Ct_df.iterrows():#dCt_df.iterrows():
    for i in range(0,len(unique_primers)):
        avg_Ct_housekeeping = mean(avg_Ct_df.loc[index, housekeeping_genes_upper])
        dCt_df.loc[index, unique_primers_analysis[i]] = avg_Ct_df.loc[index, unique_primers_analysis[i]] - avg_Ct_housekeeping


# Average control lines to calculate ddCt
dCt_df.loc['Control average'] = dCt_df.loc[control,:].mean()

# Relative expression (2^-(dCt))
rel_expression_df = pd.DataFrame()

for i in range(0, len(unique_primers)):
    for index, row in dCt_df.iterrows():
        rel_expression_df.loc[index, unique_primers_analysis[i]] = 2 ** -(dCt_df.loc[index, unique_primers_analysis[i]])

# Save relative expression values to excel
rel_expression_df.columns = original_col_names
rel_expression_df.to_excel("Data/Relative_expression_values.xlsx")

# Plot relative expression values
color_list = [''] * rel_expression_df.shape[0]
subplot_number = 1
plt.figure(figsize = (5*size_subplots,5*size_subplots))
plt.subplots_adjust(hspace = 0.5, wspace = 0.3)
# y_max = max(rel_expression_df.max()) * 1.1     # y-axis size calculated per graph

# Loop over unique primers and plot Ct values as bar graph
for i in range(0,len(unique_primers)):
    ax = plt.subplot(size_subplots, size_subplots, subplot_number)
    subplot_number += 1

    temp_values = rel_expression_df[unique_primers[i]].tolist()
    temp_samples = rel_expression_df.index
    temp_samples = temp_samples.tolist()

    for j in range(0,len(temp_samples)):
        temp_samples[j] = temp_samples[j].replace('_' + unique_primers[i], "")

    for k in range(0,len(temp_samples)):
        if temp_samples[k] in control:      # Control lines will have gray bars
            color_list[k] = 'gray'
        elif temp_samples[k] == 'Control average':
            color_list[k] = 'black'
        else:                                                                       # Investigated lines will have blue bars
            color_list[k] = 'royalblue'

    plt.bar(temp_samples, temp_values, color=color_list, alpha=0.8)
    plt.xticks(rotation='vertical')
    # plt.ylim(0,y_max)
    plt.title(unique_primers[i])

plt.savefig('Figures/Relative_expression_values.pdf', bbox_inches='tight')

#%% Calculate normalized relative gene expression ((2^-(ddCt)))

# subtract dCt of gene of interest from dCt of control average
ddCt_df = pd.DataFrame()

for i in range(0,len(unique_primers)):
    for index, row in dCt_df.iterrows():
        #ddCt_df.loc[index, unique_primers[i]] = dCt_df.loc['Control average', unique_primers[i]] - dCt_df.loc[index, unique_primers[i]]
        ddCt_df.loc[index, unique_primers_analysis[i]] = dCt_df.loc[index, unique_primers_analysis[i]] - dCt_df.loc['Control average', unique_primers_analysis[i]]

# Calculate relative ddCt (2^-2ddCt)
rel_ddCt_df = pd.DataFrame()

for index, row in ddCt_df.iterrows():
    for i in range(0,len(unique_primers)):
        rel_ddCt_df.loc[index, unique_primers_analysis[i]] = 2 ** -(ddCt_df.loc[index, unique_primers_analysis[i]])

# Save relative ddCt values to excel
rel_ddCt_df.columns = original_col_names
rel_ddCt_df.to_excel("Data/Normalized_relative_expression_values.xlsx")

## Plot relative ddCt values
color_list = [''] * rel_ddCt_df.shape[0]
subplot_number = 1
plt.figure(figsize = (5*size_subplots,5*size_subplots))
plt.subplots_adjust(hspace = 0.5, wspace = 0.3)
# y_max = max(rel_ddCt_df.max()) * 1.1     # y-axis size calculated per graph

# Loop over unique primers and plot Ct values as bar graph
for i in range(0,len(unique_primers)):
    ax = plt.subplot(size_subplots, size_subplots, subplot_number)
    subplot_number += 1

    temp_values = rel_ddCt_df[unique_primers[i]].tolist()
    temp_samples = rel_ddCt_df.index
    temp_samples = temp_samples.tolist()

    for j in range(0,len(temp_samples)):
        temp_samples[j] = temp_samples[j].replace('_' + unique_primers[i], "")

    for k in range(0,len(temp_samples)):
        if temp_samples[k] in control:      # Control lines will have gray bars
            color_list[k] = 'gray'
        elif temp_samples[k] == 'Control average':
            color_list[k] = 'black'
        else:                                                                       # Investigated lines will have blue bars
            color_list[k] = 'royalblue'

    plt.bar(temp_samples, temp_values, color=color_list, alpha=0.8)
    plt.xticks(rotation='vertical')
    # plt.ylim(0,y_max)
    plt.title(unique_primers[i])

plt.savefig('Figures/Normalized_relative_expression_values.pdf', bbox_inches='tight')

#%% Plot primer efficiency
if qPCR_system == 'LinRegPCR':
    if 'Data_output' not in user_wb.sheetnames:
        print('Error: Sheet [Data_output] not found in your input file. Make sure the sheets in your excel workbook are named correctly.\nSee the manual for more information.\n')
        sys.exit(8)

    df_output = pd.read_excel(input_file, sheet_name='Data_output')

    df_primer_eff = df_output.iloc[3:3+len(sample_names), 6]
    df_primer_eff = df_primer_eff.to_frame()
    df_primer_eff = df_primer_eff.reset_index()
    df_primer_eff['Sample'] = sample_names
    df_primer_eff['Primer'] = primer_names
    df_primer_eff['Cell_line'] = cell_lines
    del df_primer_eff['index']
    df_primer_eff = df_primer_eff.rename(columns={'Unnamed: 6' : 'Individual primer efficiency', 'Sample' : 'Sample'})
    df_primer_eff['Individual primer efficiency'] = df_primer_eff['Individual primer efficiency'].astype(float)
    
    # Remove water control from data
    for i in range(0,df_primer_eff.shape[0]):
        if df_primer_eff.loc[i,'Cell_line'] in water_controls:
            df_primer_eff = df_primer_eff.drop([i])
  
    # Make df with primersets per row
    primers_sorted = pd.DataFrame(index=range(nr_primersets),columns=range(nr_replicates * nr_samples + 1))
    row_counter = 0

    # Sort data, replicates in the same row
    for i in unique_primers:
        primers_sorted.iloc[row_counter,0] = i
        primer_df_temp = df_primer_eff[df_primer_eff['Primer'] == i]
        temp_primer_mean = primer_df_temp['Individual primer efficiency'].mean()
        primers_sorted.iloc[row_counter, 1:nr_replicates * nr_samples + 1] = primer_df_temp['Individual primer efficiency'].tolist()
        primers_sorted.loc[row_counter, 'Mean'] = temp_primer_mean
        row_counter += 1

    # Plot primer efficiency
    fig, ax = plt.subplots()
    ax.bar(unique_primers, primers_sorted['Mean'], alpha=0.5)
    ax.tick_params(axis='x', labelrotation=90)
    ax.set_ylim(0,2)
    plt.axhline(y=1.8, color='k', ls='--')

    for i in range(len(unique_primers)):
       ax.scatter([i] * nr_replicates * nr_samples, primers_sorted.iloc[i,1:nr_replicates * nr_samples + 1].values.tolist(), marker='o', c='k', s=5)

    plt.savefig('Figures/Primer_efficiency.pdf', bbox_inches='tight')

#%% Quit script
# Indicate a successful ending of the script. The web wrapper requires this, as the script can generate output
#  that is irrelevant if the script ran successfully.
sys.exit(0)
